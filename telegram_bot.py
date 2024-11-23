import time
import logging
import os
import schedule
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Poll
from telegram.ext import Updater, CommandHandler, CallbackQueryHandler, PollAnswerHandler, JobQueue
from datetime import datetime
from dotenv import load_dotenv, find_dotenv
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import Workbook, load_workbook
import csv

# Load environment variables
load_dotenv(find_dotenv())

# Timer for quiz
QUIZ_TIMEOUT_SECONDS = 15

# Function to load authorized usernames from CSV
def load_authorized_usernames(file_path):
    """Loads the list of authorized usernames from a CSV file."""
    usernames = []
    try:
        with open(file_path, mode="r", encoding="utf-8") as file:
            reader = csv.DictReader(file)
            for row in reader:
                if "Telegram Username" in row:  # Check if the column exists
                    usernames.append(row["Telegram Username"])
    except FileNotFoundError:
        logging.warning(f"⚠️ File {file_path} not found. Authorized user list will be empty.")
    except Exception as e:
        logging.error(f"⚠️ Error reading file {file_path}: {e}")
    return usernames

# Configuration for users and admins
csv_file_path = "registration_log.csv"  # Specify the correct path
authorized_usernames = load_authorized_usernames(csv_file_path)
SUPERADMIN_USERNAME = "Alexander_Dym"
file_path = "updated_bot_list.xlsx"

# Initialize the Excel file
def initialize_excel():
    """Creates Excel file with sheets for each quiz day if it doesn't exist."""
    if not os.path.exists(file_path):
        wb = Workbook()
        for i in range(1, 8):  # Create sheets "Day 1", "Day 2", ..., "Day 7"
            sheet = wb.create_sheet(title=f"Day {i}")
            headers = [
                "User ID", "Username", "First Name", "Last Name", "First Interaction Date", 
                "Quiz Participation Date", f"Day {i} Answer", "Last Interaction Date"
            ]
            sheet.append(headers)
        wb.remove(wb["Sheet"])  # Remove default sheet
        wb.save(file_path)

# Class for quiz questions
class QuizQuestion:
    def __init__(self, question="", answers=None, correct_answer=""):
        self.question = question
        self.answers = answers if answers is not None else []
        self.correct_answer = correct_answer
        self.correct_answer_position = self.__get_correct_answer_position__()

    def __get_correct_answer_position__(self):
        for index, answer in enumerate(self.answers):
            if answer.strip().lower() == self.correct_answer.strip().lower():
                return index
        return -1

# Quiz questions for 7 days
quiz_questions = [
    QuizQuestion("Which of these is a popular affiliate marketing model?", ["Cost per click (CPC)", "Cost per lead (CPL)", "Pay per hire (PPH)"], "Cost per lead (CPL)"),
    QuizQuestion("Какая столица Франции?", ["Лондон", "Берлин", "Париж"], "Париж"),
    QuizQuestion("Сколько планет в Солнечной системе?", ["7", "8", "9"], "8"),
    QuizQuestion("Какое самое большое животное на земле?", ["Слон", "Синий кит", "Жираф"], "Синий кит"),
    QuizQuestion("Сколько секунд в одной минуте?", ["60", "100", "120"], "60"),
    QuizQuestion("Какой элемент обозначается символом O?", ["Кислород", "Водород", "Азот"], "Кислород"),
    QuizQuestion("Кто является создателем теории относительности?", ["Ньютон", "Эйнштейн", "Галилей"], "Эйнштейн"),
]

# Times for each quiz day
daily_times = ["13:21", "13:22", "13:23", "13:28", "21:19", "22:00", "10:00"]

# Record user response in Excel
def record_user_response(user_id, username, first_name, last_name, day, selected_answer, correct_answer, result):
    """Records user response in the corresponding sheet for each day."""
    wb = load_workbook(file_path)
    sheet_name = f"Day {day + 1}"
    
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)
    sheet = wb[sheet_name]

    # Green for "Correct", Red for "Incorrect"
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Formatting result and filling color
    first_interaction_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    result_text = "Верно" if result else "Неверно"
    result_fill = green_fill if result else red_fill

    # Search or create user entry
    user_found = False
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
        if row[0].value == user_id:
            row[6].value = result_text
            row[6].fill = result_fill
            row[7].value = first_interaction_date  # Update last interaction date
            user_found = True
            break

    if not user_found:
        new_entry = [
            user_id, username, first_name, last_name, first_interaction_date, first_interaction_date,
            result_text, first_interaction_date
        ]
        sheet.append(new_entry)
        sheet.cell(row=sheet.max_row, column=7).fill = result_fill

    wb.save(file_path)
    logging.info(f"Result for user {user_id} recorded on {sheet_name}: {result_text}")

# Command for superadmin to get the results file
def list_handler(update, context):
    user = update.message.from_user

    if user.username == SUPERADMIN_USERNAME:
        try:
            with open(file_path, 'rb') as file:
                context.bot.send_document(chat_id=update.effective_chat.id, document=file, filename="quiz_results.xlsx")
                update.message.reply_text("Here are the current quiz results.")
        except Exception as e:
            update.message.reply_text(f"Failed to send the file: {str(e)}")
    else:
        update.message.reply_text("⛔ You don't have access to this command.")

# Command to start the quiz for the user
def start_command_handler(update, context):
    if not is_authorized_user(update):
        update.message.reply_text("⛔ Извините, доступ к этому боту ограничен.")
        return
    
    chat_id = update.effective_chat.id
    user_data = context.user_data.setdefault(chat_id, {})
    user_data['waiting_for_answer'] = False
    
    # Отправляем приветственное сообщение с изображением
    image_url = "https://mailer.ucliq.com/wizz/frontend/assets/files/customer/kd629xy3hj208/img/girl-wearing-santa-s-hat-%281%29.jpg"
    welcome_text = (
        "🎄🎅 Добро пожаловать на наш праздничный розыгрыш от 17 по 24 декабря 🎅🎄\n\n"
        "✨ Мы будем задавать вам вопросы каждый день и разыгрывать призы среди участников\n"
        "🎁 Главный приз — уникальный подарок в канун Рождества\n\n"
        "Нажмите кнопку ниже, чтобы участвовать в викторине и выиграть 🎉"
    )
    context.bot.send_photo(chat_id=chat_id, photo=image_url, caption=welcome_text)

    keyboard = [[InlineKeyboardButton("📝 Участвовать в викторине", callback_data="participate")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    sent_message = context.bot.send_message(
        chat_id=chat_id,
        text="Нажмите 'Участвовать в викторине', чтобы начать.",
        reply_markup=reply_markup
    )
    user_data['participate_message_id'] = sent_message.message_id

# Callback for participating in quiz
def participate_handler(update, context):
    if not is_authorized_user(update):
        update.callback_query.answer("⛔ Доступ запрещен.")
        return

    query = update.callback_query
    query.answer()
    
    chat_id = query.message.chat_id
    user_data = context.user_data.setdefault(chat_id, {})

    if user_data.get('waiting_for_answer', False):
        return

    if 'participate_message_id' in user_data:
        context.bot.delete_message(chat_id=chat_id, message_id=user_data['participate_message_id'])
        del user_data['participate_message_id']

    send_daily_quiz(context, context.bot_data.get('current_day', 0))
    user_data['waiting_for_answer'] = True

# Function to send quiz question
def send_daily_quiz(context, day):
    if day < len(quiz_questions):
        question = quiz_questions[day]
        logging.info(f"Sending question for day {day + 1}: '{question.question}'")
        
        for chat_id, user_data in context.dispatcher.user_data.items():
            if user_data.get('waiting_for_answer', False):
                continue  # Skip users already waiting for an answer
            add_quiz_question(context, question, chat_id, day)
            user_data['waiting_for_answer'] = True
    else:
        logging.error(f"Day {day + 1} is out of range for questions.")

# Function to send quiz question to user
def add_quiz_question(context, quiz_question, chat_id, day):
    message = context.bot.send_poll(
        chat_id=chat_id,
        question=quiz_question.question,
        options=quiz_question.answers,
        type=Poll.QUIZ,
        correct_option_id=quiz_question.correct_answer_position,
        open_period=QUIZ_TIMEOUT_SECONDS,
        is_anonymous=False,
        explanation="Попробуйте ещё раз"
    )
    context.bot_data.update({message.poll.id: {'chat_id': message.chat.id, 'day': day}})

    # Start timer for timeout check
    context.job_queue.run_once(handle_timeout, QUIZ_TIMEOUT_SECONDS + 1, context={"chat_id": chat_id, "day": day})

# Handle timeout
def handle_timeout(context):
    job_context = context.job.context
    chat_id = job_context["chat_id"]
    day = job_context["day"]

    user_data = context.dispatcher.user_data.get(chat_id, {})
    if user_data.get('waiting_for_answer', True):
        context.bot.send_message(chat_id=chat_id, text="❌ Время вышло! Вы не успели ответить. Попробуйте ещё раз завтра.")
        # Record result as "incorrect"
        user = context.bot.get_chat(chat_id)
        record_user_response(
            user_id=user.id,
            username=user.username,
            first_name=user.first_name,
            last_name=user.last_name,
            day=day,
            selected_answer="",
            correct_answer="",
            result=False
        )
        user_data['waiting_for_answer'] = False

# Poll answer handler
def poll_handler(update, context):
    poll_answer = update.poll_answer
    user_id = poll_answer.user.id
    selected_option_id = poll_answer.option_ids[0]  # ID of the option selected by the user
    
    # Determine current quiz day from context
    poll_data = context.bot_data.get(poll_answer.poll_id, {})
    day = poll_data.get('day', 0)
    question = quiz_questions[day]
    correct_option_id = question.correct_answer_position
    is_correct = (selected_option_id == correct_option_id)  # Check if answer is correct
    
    # Record result in Excel
    record_user_response(
        user_id=user_id,
        username=poll_answer.user.username,
        first_name=poll_answer.user.first_name,
        last_name=poll_answer.user.last_name,
        day=day,
        selected_answer=question.answers[selected_option_id],
        correct_answer=question.answers[correct_option_id],
        result=is_correct
    )

    # Respond to user
    if is_correct:
        context.bot.send_message(chat_id=user_id, text="🎉 Поздравляем! 🎉\n\nВы выиграли свой приз! 🏆✨ Свяжитесь со своим менеджером, чтобы получить его. Удачи в следующих вопросах и до новых побед! 💫")
    else:
        context.bot.send_message(chat_id=user_id, text="❌ Упс, это неправильный ответ! Но не сдавайтесь! 🎯\n\nПопробуйте ещё раз завтра и приближайтесь к своему призу! 🏅")

    # Mark that user answered
    context.dispatcher.user_data[user_id]['waiting_for_answer'] = False

# Check if user is authorized
def is_authorized_user(update):
    user = update.effective_user
    return user.username == SUPERADMIN_USERNAME or user.username in authorized_usernames

# Main function
def main():
    initialize_excel()
    updater = Updater(DefaultConfig.TELEGRAM_TOKEN, use_context=True)
    dp = updater.dispatcher
    dp.add_handler(CommandHandler("start", start_command_handler))
    dp.add_handler(CommandHandler("list", list_handler))  # Command for superadmin to get the results list
    dp.add_handler(CallbackQueryHandler(participate_handler, pattern="participate"))
    dp.add_handler(PollAnswerHandler(poll_handler))

    # Schedule daily quiz questions
    job_queue = updater.job_queue
    for i, daily_time in enumerate(daily_times):
        schedule.every().day.at(daily_time).do(lambda day=i: job_queue.run_once(lambda _: send_daily_quiz(dp.context, day), 0))

    # Start polling and scheduling
    updater.start_polling()
    logging.info("Start polling mode")
    
    while True:
        schedule.run_pending()
        time.sleep(1)

# Default configuration
class DefaultConfig:
    PORT = int(os.environ.get("PORT", 3978))
    TELEGRAM_TOKEN = "7603983242:AAGYo--n9YxQlhiJOwydp3HorHedHAwZtlc"
    MODE = os.environ.get("MODE", "polling")
    WEBHOOK_URL = os.environ.get("WEBHOOK_URL", "")

if __name__ == "__main__":
    logging.basicConfig(format="%(asctime)s - %(levelname)s - %(message)s", level=logging.INFO)
    main()
