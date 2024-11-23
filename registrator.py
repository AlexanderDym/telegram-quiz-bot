import logging
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Updater, CommandHandler, MessageHandler, Filters, ConversationHandler, CallbackContext, CallbackQueryHandler
from datetime import datetime

# Путь к таблице
allowed_users_file = "user_list.xlsx"  # Excel-файл с разрешёнными пользователями
registration_log_file = "registration_log.csv"  # CSV-файл для регистрации
main_bot_link = "https://t.me/tramee_chrismass_bot"  # Ссылка на основного бота

# Состояния для ConversationHandler
ASK_TRAFFEE_USERNAME = range(1)

# Проверка наличия пользователя в таблице
def is_user_in_list(username):
    """Проверяет, есть ли пользователь в таблице user_list.xlsx и зарегистрирован ли он"""
    if not os.path.exists(allowed_users_file):
        logging.error(f"Файл {allowed_users_file} отсутствует.")
        return False, False

    try:
        df = pd.read_excel(allowed_users_file)
        if "username" not in df.columns:
            logging.error("В файле отсутствует столбец 'username'.")
            return False, False

        # Приведение введённого имени и таблицы к нижнему регистру
        username_cleaned = username.strip().lower()
        user_list = df["username"].dropna().astype(str).str.strip().str.lower().tolist()

        if username_cleaned in user_list:
            wb = load_workbook(allowed_users_file)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                cell_value = str(row[0].value).strip().lower()
                if cell_value == username_cleaned:
                    # Проверка цвета ячейки (если пользователь уже зарегистрирован)
                    if row[0].fill.start_color.index == "00FF00":
                        return True, True  # Найден и уже зарегистрирован
                    return True, False  # Найден, но ещё не зарегистрирован

        return False, False
    except Exception as e:
        logging.error(f"Ошибка при проверке наличия пользователя: {e}")
        return False, False

# Выделение пользователя зелёным цветом
def mark_user_as_registered(username):
    """Отмечает пользователя зелёным цветом в таблице user_list.xlsx"""
    if not os.path.exists(allowed_users_file):
        logging.error(f"Файл {allowed_users_file} отсутствует. Невозможно отметить пользователя.")
        return
    
    try:
        wb = load_workbook(allowed_users_file)
        ws = wb.active

        # Поиск строки с указанным username
        username_cleaned = username.strip().lower()
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            cell_value = str(row[0].value).strip().lower()
            if cell_value == username_cleaned:
                green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                for cell in row:
                    cell.fill = green_fill
                wb.save(allowed_users_file)  # Сохраняем изменения
                logging.info(f"Пользователь {username} успешно отмечен зелёным.")
                return

        logging.warning(f"Пользователь {username} не найден в таблице.")
    except Exception as e:
        logging.error(f"Ошибка при сохранении регистрации пользователя {username}: {e}")

# Логирование регистрации в CSV
def log_registration(username, telegram_username):
    """Записывает данные регистрации в CSV-файл"""
    data = {
        "Trafee Username": username,
        "Telegram Username": telegram_username,
        "Registration Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    file_exists = os.path.exists(registration_log_file)

    try:
        with open(registration_log_file, mode="a", encoding="utf-8") as file:
            if not file_exists:
                file.write("Trafee Username,Telegram Username,Registration Date\n")  # Заголовок
            file.write(f"{data['Trafee Username']},{data['Telegram Username']},{data['Registration Date']}\n")
        logging.info(f"Регистрация пользователя {username} записана в CSV.")
    except Exception as e:
        logging.error(f"Ошибка при записи данных в CSV: {e}")

# Функция для старта регистрации
def start(update: Update, context: CallbackContext):
    update.message.reply_text(
        "👋 Добро пожаловать!\n\n"
        "Пожалуйста, введите ваш *Trafee username* для проверки доступа."
    )
    return ASK_TRAFFEE_USERNAME

# Проверка username
def check_username(update: Update, context: CallbackContext):
    username = update.message.text.strip()
    telegram_username = update.message.from_user.username

    found, registered = is_user_in_list(username)
    if found and registered:
        # Если пользователь уже зарегистрирован
        update.message.reply_text(
            f"✅ {username}, вы уже зарегистрированы!\n"
            f"Напоминаем ссылку на [основного бота]({main_bot_link}).",
            parse_mode="Markdown"
        )
        return ConversationHandler.END
    elif found:
        # Если пользователь найден, но ещё не зарегистрирован
        mark_user_as_registered(username)
        log_registration(username, telegram_username)  # Логируем регистрацию
        update.message.reply_text(
            f"🎉 Поздравляем, {username}! Вы успешно зарегистрированы.\n"
            f"Перейдите в [основного бота]({main_bot_link}), чтобы продолжить участие.",
            parse_mode="Markdown"
        )
        return ConversationHandler.END
    else:
        # Если пользователь не найден
        keyboard = [[InlineKeyboardButton("🔄 Попробовать ещё раз", callback_data="retry")]]
        reply_markup = InlineKeyboardMarkup(keyboard)

        update.message.reply_text(
            "😔 Извините, но ваш Trafee username отсутствует в списке разрешённых пользователей.\n"
            "Если вы считаете, что это ошибка, свяжитесь с вашим менеджером.",
            reply_markup=reply_markup
        )
        return ASK_TRAFFEE_USERNAME

# Обработка кнопки "Попробовать ещё раз"
def retry_handler(update: Update, context: CallbackContext):
    query = update.callback_query
    query.answer()
    query.edit_message_text("Пожалуйста, введите ваш *Trafee username* ещё раз.")
    return ASK_TRAFFEE_USERNAME

# Отмена регистрации
def cancel(update: Update, context: CallbackContext):
    update.message.reply_text("🚫 Регистрация отменена. Если захотите попробовать снова, напишите /start.")
    return ConversationHandler.END

# Команда /user_list для отправки таблицы
def send_user_list(update: Update, context: CallbackContext):
    if os.path.exists(allowed_users_file):
        update.message.reply_document(
            document=open(allowed_users_file, "rb"),
            filename="user_list.xlsx",
            caption="📋 Вот актуальный список пользователей."
        )
    else:
        update.message.reply_text("❌ Таблица с пользователями отсутствует.")

# Основная функция
def main():
    updater = Updater("8150438145:AAE47-REyY4_7_3IpYbazVI_Lw4GAGSmWec", use_context=True)
    dp = updater.dispatcher

    # ConversationHandler для регистрации
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            ASK_TRAFFEE_USERNAME: [
                MessageHandler(Filters.text & ~Filters.command, check_username),
                CallbackQueryHandler(retry_handler, pattern="^retry$")
            ],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    dp.add_handler(conv_handler)

    # Команда для отправки таблицы
    dp.add_handler(CommandHandler("user_list", send_user_list))

    # Запуск бота
    updater.start_polling()
    logging.info("Регистрационный бот запущен")
    updater.idle()

if __name__ == "__main__":
    logging.basicConfig(format="%(asctime)s - %(levelname)s - %(message)s", level=logging.INFO)
    main()
